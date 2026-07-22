# -*- coding: utf-8 -*-
"""
Moteur de génération des plannings de formation (IMSR - ECF-CERCA)
Reproduit le modèle "Planning_contrat_pro_TP_ECSR" (calendrier annuel Sept->Août,
codes couleur Centre / FOAD / Entreprise / Examens, jours fériés automatiques)
et propose un format compact pour les formations courtes.
"""

from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Couleurs reprises du modèle fourni
# ---------------------------------------------------------------------------
COLOR_CENTRE = "92D050"
COLOR_FOAD = "F1A983"
COLOR_ENTREPRISE = "61CBF3"
COLOR_EXAMENS = None  # pas de remplissage dans le modèle, texte en gras
COLOR_FERIE = "ACB9CA"
COLOR_HEADER_FONT = "002060"
COLOR_TITLE_FILL = "1B2A4A"   # navy ECF-CERCA (utilisé pour les titres)
COLOR_ACCENT = "0F7B7B"       # teal ECF-CERCA

MODALITES = ["Centre", "FOAD", "Entreprise", "Examens"]
JOURS_LETTRE = ["L", "M", "M", "J", "V", "S", "D"]  # Lundi..Dimanche
MOIS_LABELS = ["Septembre", "Octobre", "Novembre", "Décembre", "Janvier",
               "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août"]

THIN = Side(style="thin", color="BFBFBF")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Jours fériés français (fixes + mobiles, calcul de Pâques - algorithme de Meeus)
# ---------------------------------------------------------------------------
def easter_sunday(year: int) -> date:
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def french_holidays(year: int) -> dict:
    """Retourne les jours fériés français pour une année civile donnée."""
    easter = easter_sunday(year)
    holidays = {
        date(year, 1, 1): "Jour de l'an",
        date(year, 5, 1): "Fête du travail",
        date(year, 5, 8): "Victoire 1945",
        date(year, 7, 14): "Fête nationale",
        date(year, 8, 15): "Assomption",
        date(year, 11, 1): "Toussaint",
        date(year, 11, 11): "Armistice 1918",
        date(year, 12, 25): "Noël",
        easter + timedelta(days=1): "L. de Pâques",
        easter + timedelta(days=39): "Ascension",
        easter + timedelta(days=50): "L. Pentecôte",
    }
    return holidays


def holidays_for_range(start: date, end: date) -> dict:
    result = {}
    for y in range(start.year, end.year + 1):
        result.update(french_holidays(y))
    return {d: label for d, label in result.items() if start <= d <= end}


# ---------------------------------------------------------------------------
# Utilitaires période / modalité
# ---------------------------------------------------------------------------
def build_day_modality_map(periods: list, start: date, end: date) -> dict:
    """
    periods: liste de dicts {"debut": date, "fin": date, "modalite": str}
    Retourne {date: modalite} pour tous les jours ouvrés (Lun-Ven) couverts.
    En cas de chevauchement, la période la plus tardive dans la liste l'emporte.
    """
    day_map = {}
    for p in periods:
        d0, d1, mod = p["debut"], p["fin"], p["modalite"]
        if not d0 or not d1 or not mod:
            continue
        d0 = max(d0, start)
        d1 = min(d1, end)
        cur = d0
        while cur <= d1:
            if cur.weekday() < 5:  # jours ouvrés uniquement (Lun-Ven)
                day_map[cur] = mod
            cur += timedelta(days=1)
    return day_map


JOURS_SEMAINE = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]


def apply_directives(day_map: dict, directives: list, start: date, end: date):
    """
    Applique des directives récurrentes par jour de semaine (ex: "CCP2 en FOAD
    tous les jeudis du 5/01 au 20/03") par-dessus la répartition par périodes.
    Les directives sont prioritaires sur les périodes de base.

    directives: liste de dicts {"jour": "Jeudi", "modalite": str,
                                 "module": str|None, "debut": date, "fin": date}

    Retourne (day_map mis à jour, day_module_map {date: code_module}).
    """
    day_module_map = {}
    for d in directives:
        jour = d.get("jour")
        mod = d.get("modalite")
        if not jour or not mod or jour not in JOURS_SEMAINE:
            continue
        weekday_idx = JOURS_SEMAINE.index(jour)
        d0 = max(d.get("debut") or start, start)
        d1 = min(d.get("fin") or end, end)
        if d0 > d1:
            continue
        # se positionner sur la première occurrence du jour de semaine visé
        offset = (weekday_idx - d0.weekday()) % 7
        cur = d0 + timedelta(days=offset)
        while cur <= d1:
            day_map[cur] = mod
            module_code = d.get("module")
            if module_code:
                day_module_map[cur] = module_code
            cur += timedelta(days=7)
    return day_map, day_module_map


def planned_hours_by_module(day_module_map: dict) -> dict:
    """Retourne {code_module: heures planifiées} à partir des directives récurrentes
    (1 jour = 7h, cohérent avec le reste du calcul des heures du modèle)."""
    totals = {}
    for _, code in day_module_map.items():
        totals[code] = totals.get(code, 0) + 7
    return totals


def academic_years_covered(start: date, end: date):
    """Retourne la liste des années de début d'année scolaire (sept N -> août N+1)
    couvertes par [start, end]."""
    years = []
    y = start.year if start.month >= 9 else start.year - 1
    cur_start = date(y, 9, 1)
    while cur_start <= end:
        cur_end = date(y + 1, 8, 31)
        if cur_end >= start:
            years.append(y)
        y += 1
        cur_start = date(y, 9, 1)
    return years


# ---------------------------------------------------------------------------
# Styles communs
# ---------------------------------------------------------------------------
def _fill(hex_color):
    if not hex_color:
        return PatternFill(fill_type=None)
    return PatternFill(fill_type="solid", fgColor="FF" + hex_color, bgColor="FF" + hex_color)


def _title_font(size=16, bold=True, color="FFFFFF"):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _header_font(size=11, bold=True, color=COLOR_HEADER_FONT):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _day_font(size=9, bold=False, color=COLOR_HEADER_FONT):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def modality_fill(modality_or_label, is_holiday=False):
    if is_holiday:
        return _fill(COLOR_FERIE)
    return {
        "Centre": _fill(COLOR_CENTRE),
        "FOAD": _fill(COLOR_FOAD),
        "Entreprise": _fill(COLOR_ENTREPRISE),
        "Examens": _fill(COLOR_EXAMENS),
    }.get(modality_or_label, _fill(None))


# ---------------------------------------------------------------------------
# Génération d'une feuille "année scolaire" (grand format, formations longues)
# ---------------------------------------------------------------------------
def write_year_sheet(wb, sheet_name, year_start, learner_name, contract_type,
                      training_type, day_map, holidays):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Titres
    ws.merge_cells("B2:S2")
    ws["B2"] = f"{learner_name or 'Prénom NOM'}"
    ws["B2"].font = _header_font(size=14)
    ws.merge_cells("T2:AK2")
    ws["T2"] = training_type
    ws["T2"].font = _header_font(size=16)
    ws["T2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("B4:P4")
    ws["B4"] = f"Calendrier {year_start} - {year_start + 1}"
    ws["B4"].font = _header_font(size=14)
    ws["B4"].alignment = Alignment(horizontal="right")

    ws.merge_cells("T4:AK4")
    ws["T4"] = contract_type
    ws["T4"].font = _header_font(size=14)
    ws["T4"].alignment = Alignment(horizontal="center")

    # 12 blocs de mois, 3 colonnes chacun, à partir de la colonne B
    month_starts = []
    col = 2  # B
    month_dates = []
    y = year_start
    m = 9
    for i in range(12):
        month_dates.append(date(y, m, 1))
        m += 1
        if m == 13:
            m = 1
            y += 1

    for i, mstart in enumerate(month_dates):
        c0 = col + i * 3
        c0_letter = get_column_letter(c0)
        c2_letter = get_column_letter(c0 + 2)
        ws.merge_cells(f"{c0_letter}6:{c2_letter}6")
        cell = ws[f"{c0_letter}6"]
        cell.value = MOIS_LABELS[(mstart.month - 9) % 12]
        cell.font = _header_font(size=11)
        cell.alignment = Alignment(horizontal="center")
        month_starts.append((c0, mstart))

    # colonnes étroites
    for i in range(2, 2 + 12 * 3):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 3.2 if (i - 2) % 3 != 2 else 9.5

    # Grille des jours (lignes 7 à 37 -> jours 1 à 31)
    for c0, mstart in month_starts:
        y_, m_ = mstart.year, mstart.month
        if m_ == 12:
            next_month = date(y_ + 1, 1, 1)
        else:
            next_month = date(y_, m_ + 1, 1)
        nb_days = (next_month - mstart).days

        for d in range(1, nb_days + 1):
            cur = date(y_, m_, d)
            r = 7 + d - 1
            col_num = get_column_letter(c0)
            col_letter = get_column_letter(c0)
            col_letter_dow = get_column_letter(c0 + 1)
            col_letter_act = get_column_letter(c0 + 2)

            c_num = ws[f"{col_letter}{r}"]
            c_num.value = d
            c_num.font = _day_font()
            c_num.alignment = Alignment(horizontal="center")
            c_num.border = BORDER_THIN

            c_dow = ws[f"{col_letter_dow}{r}"]
            c_dow.value = JOURS_LETTRE[cur.weekday()]
            c_dow.font = _day_font()
            c_dow.alignment = Alignment(horizontal="center")
            c_dow.border = BORDER_THIN

            c_act = ws[f"{col_letter_act}{r}"]
            is_holiday = cur in holidays
            if is_holiday:
                c_act.value = holidays[cur]
                c_act.fill = modality_fill(None, is_holiday=True)
            elif cur in day_map:
                mod = day_map[cur]
                c_act.value = mod
                c_act.fill = modality_fill(mod)
                if mod == "Examens":
                    c_act.font = Font(name="Calibri", size=8, bold=True, color=COLOR_HEADER_FONT)
            c_act.alignment = Alignment(horizontal="center")
            c_act.border = BORDER_THIN

    # Ligne 38 : total jours du mois avec activité (formule COUNTIF sur la colonne "activité")
    for c0, mstart in month_starts:
        act_col = get_column_letter(c0 + 2)
        first_row, last_row = 7, 37
        cell = ws[f"{act_col}38"]
        cell.value = (f'=COUNTIF({act_col}{first_row}:{act_col}{last_row},"Centre")'
                      f'+COUNTIF({act_col}{first_row}:{act_col}{last_row},"FOAD")'
                      f'+COUNTIF({act_col}{first_row}:{act_col}{last_row},"Entreprise")'
                      f'+COUNTIF({act_col}{first_row}:{act_col}{last_row},"Examens")')
        cell.font = _day_font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Récapitulatif heures (colonnes AM/AN/AO), sur toute la plage d'activité
    act_ranges = ",".join(
        f'{get_column_letter(c0 + 2)}7:{get_column_letter(c0 + 2)}37' for c0, _ in month_starts
    )
    recap_row0 = 6
    labels = ["Centre", "FOAD", "Examens", "Entreprise"]
    ws["AM6"] = "RÉCAPITULATIF HEURES"
    ws["AM6"].font = _header_font(size=11)
    for idx, label in enumerate(labels):
        r = recap_row0 + 1 + idx
        ws[f"AM{r}"] = label
        ws[f"AM{r}"].font = _day_font(bold=True)
        formula_parts = [f'COUNTIF({get_column_letter(c0+2)}7:{get_column_letter(c0+2)}37,"{label}")'
                          for c0, _ in month_starts]
        ws[f"AN{r}"] = "=(" + "+".join(formula_parts) + ")*7"
        ws[f"AO{r}"] = "h"
    total_row = recap_row0 + 1 + len(labels)
    ws[f"AM{total_row}"] = "TOTAL"
    ws[f"AM{total_row}"].font = _day_font(bold=True)
    ws[f"AN{total_row}"] = f"=SUM(AN{recap_row0+1}:AN{recap_row0+len(labels)})"
    ws[f"AO{total_row}"] = "h"

    # Zone commentaires
    ws.merge_cells(f"B{total_row+2}:AK{total_row+2}")
    ws[f"B{total_row+2}"] = "COMMENTAIRES :"
    ws[f"B{total_row+2}"].font = _day_font(bold=True)
    ws.merge_cells(f"B{total_row+3}:AK{total_row+5}")

    return ws


# ---------------------------------------------------------------------------
# Génération d'une feuille compacte (formations courtes)
# ---------------------------------------------------------------------------
def write_short_sheet(wb, sheet_name, learner_name, contract_type, training_type,
                       start, end, day_map, holidays):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = learner_name or "Prénom NOM"
    ws["A1"].font = _header_font(size=14)

    ws.merge_cells("A2:E2")
    ws["A2"] = training_type
    ws["A2"].font = _header_font(size=13)

    ws.merge_cells("A3:E3")
    ws["A3"] = contract_type
    ws["A3"].font = _day_font(size=10)

    headers = ["Date", "Jour", "Modalité / Observation"]
    for i, h in enumerate(headers):
        c = ws.cell(row=5, column=1 + i, value=h)
        c.font = _header_font()
        c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 28

    row = 6
    cur = start
    while cur <= end:
        ws.cell(row=row, column=1, value=cur.strftime("%d/%m/%Y")).alignment = Alignment(horizontal="center")
        dow_full = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"][cur.weekday()]
        ws.cell(row=row, column=2, value=dow_full).alignment = Alignment(horizontal="center")
        c_act = ws.cell(row=row, column=3)
        is_holiday = cur in holidays
        if is_holiday:
            c_act.value = holidays[cur]
            c_act.fill = modality_fill(None, is_holiday=True)
        elif cur in day_map:
            mod = day_map[cur]
            c_act.value = mod
            c_act.fill = modality_fill(mod)
        c_act.alignment = Alignment(horizontal="center")
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = BORDER_THIN
            ws.cell(row=row, column=col).font = _day_font()
        row += 1
        cur += timedelta(days=1)

    # Récap heures
    last_row = row - 1
    ws.cell(row=row + 1, column=1, value="RÉCAPITULATIF HEURES").font = _header_font()
    labels = ["Centre", "FOAD", "Examens", "Entreprise"]
    r0 = row + 2
    for idx, label in enumerate(labels):
        r = r0 + idx
        ws.cell(row=r, column=1, value=label).font = _day_font(bold=True)
        ws.cell(row=r, column=2, value=f'=COUNTIF(C6:C{last_row},"{label}")*7')
        ws.cell(row=r, column=3, value="h")
    r_total = r0 + len(labels)
    ws.cell(row=r_total, column=1, value="TOTAL").font = _day_font(bold=True)
    ws.cell(row=r_total, column=2, value=f"=SUM(B{r0}:B{r0+len(labels)-1})")
    ws.cell(row=r_total, column=3, value="h")

    ws.cell(row=r_total + 2, column=1, value="COMMENTAIRES :").font = _day_font(bold=True)

    return ws


# ---------------------------------------------------------------------------
# Feuille Modules (CCP / BC + volumes horaires)
# ---------------------------------------------------------------------------
def write_modules_sheet(wb, modules: list, learner_name, training_type, planned_hours: dict = None):
    planned_hours = planned_hours or {}
    ws = wb.create_sheet("Modules")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Modules — {training_type}" + (f" — {learner_name}" if learner_name else "")
    ws["A1"].font = _header_font(size=13)

    headers = ["Code (CCP / BC / UC)", "Intitulé du module", "Modalité",
               "Heures de formation", "Heures de stage", "Total prévu (h)",
               "Heures planifiées (directives)", "Écart (h)"]
    for i, h in enumerate(headers):
        c = ws.cell(row=3, column=1 + i, value=h)
        c.fill = _fill(COLOR_ACCENT)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    widths = {"A": 18, "B": 40, "C": 14, "D": 16, "E": 14, "F": 14, "G": 22, "H": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    row = 4
    for mod in modules:
        code = mod.get("code", "")
        intitule = mod.get("intitule", "")
        modalite = mod.get("modalite", "")
        h_formation = mod.get("heures_formation", 0)
        h_stage = mod.get("heures_stage", 0)
        if not code and not intitule:
            continue
        ws.cell(row=row, column=1, value=code).border = BORDER_THIN
        ws.cell(row=row, column=2, value=intitule).border = BORDER_THIN
        c_mod = ws.cell(row=row, column=3, value=modalite)
        c_mod.border = BORDER_THIN
        c_mod.alignment = Alignment(horizontal="center")
        c_hf = ws.cell(row=row, column=4, value=h_formation)
        c_hf.border = BORDER_THIN
        c_hf.alignment = Alignment(horizontal="center")
        c_hs = ws.cell(row=row, column=5, value=h_stage)
        c_hs.border = BORDER_THIN
        c_hs.alignment = Alignment(horizontal="center")
        c_tot = ws.cell(row=row, column=6, value=f"=D{row}+E{row}")
        c_tot.border = BORDER_THIN
        c_tot.alignment = Alignment(horizontal="center")

        planned = planned_hours.get(code, 0)
        c_plan = ws.cell(row=row, column=7, value=planned)
        c_plan.border = BORDER_THIN
        c_plan.alignment = Alignment(horizontal="center")

        c_ecart = ws.cell(row=row, column=8, value=f"=F{row}-G{row}")
        c_ecart.border = BORDER_THIN
        c_ecart.alignment = Alignment(horizontal="center")
        row += 1

    last_data_row = row - 1
    ws.cell(row=row + 1, column=1, value="TOTAL").font = _day_font(bold=True)
    if last_data_row >= 4:
        for col in (4, 5, 6, 7):
            col_letter = get_column_letter(col)
            ws.cell(row=row + 1, column=col, value=f"=SUM({col_letter}4:{col_letter}{last_data_row})")
    else:
        for col in (4, 5, 6, 7):
            ws.cell(row=row + 1, column=col, value=0)
    for col in (4, 5, 6, 7):
        ws.cell(row=row + 1, column=col).font = _day_font(bold=True)

    if planned_hours:
        ws.cell(row=row + 3, column=1,
                value="« Heures planifiées » = jours issus des directives récurrentes de "
                      "l'onglet 3 (jour de semaine × modalité × module) sur la période du planning.")
        ws.cell(row=row + 3, column=1).font = Font(name="Calibri", size=9, italic=True, color="7F7F7F")

    return ws


# ---------------------------------------------------------------------------
# Fonction principale
# ---------------------------------------------------------------------------
def generate_workbook(learner_name, contract_type, training_type, mode,
                       start, end, periods, modules, directives=None):
    """
    mode: "long" (calendrier annuel Sept->Août, une feuille par année scolaire)
          "short" (feuille compacte liste de dates)
    periods: liste de dicts {"debut": date, "fin": date, "modalite": str}
    modules: liste de dicts {"code": str, "intitule": str, "volume": float}
    directives: liste de dicts {"jour": "Jeudi", "modalite": str, "module": str|None,
                                 "debut": date, "fin": date}
                Directives récurrentes (ex: "CCP2 en FOAD tous les jeudis"),
                prioritaires sur les périodes de base.
    """
    directives = directives or []
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    holidays = holidays_for_range(start, end)
    day_map = build_day_modality_map(periods, start, end)
    day_map, day_module_map = apply_directives(day_map, directives, start, end)
    # Les jours fériés restent prioritaires même sur une directive récurrente
    for d in list(day_map.keys()):
        if d in holidays:
            day_map.pop(d, None)
            day_module_map.pop(d, None)
    planned_hours = planned_hours_by_module(day_module_map)

    if mode == "long":
        for y in academic_years_covered(start, end):
            sheet_name = f"{y} - {y+1}"
            write_year_sheet(wb, sheet_name, y, learner_name, contract_type,
                              training_type, day_map, holidays)
    else:
        write_short_sheet(wb, "Planning", learner_name, contract_type, training_type,
                           start, end, day_map, holidays)

    write_modules_sheet(wb, modules, learner_name, training_type, planned_hours)

    return wb
